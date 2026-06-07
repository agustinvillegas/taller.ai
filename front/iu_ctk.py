import customtkinter as ctk
import threading
import json
import os
from datetime import datetime
from back.ai import analizar_datos_web
from back.mail import enviar_mail
from back.sesion import cargar_sesion, guardar_sesion, cerrar_sesion, actualizar_empresa_sesion
from back.empresa import CAMPOS_EMPRESA, empresa_completa, get_contexto_empresa, get_encabezado_empresa
import re
import PIL

import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")
ctk.set_widget_scaling(1.12)
ctk.set_window_scaling(1.0)

COLORS = {
    "bg":          "#111318",
    "surface":     "#181B22",
    "surface2":    "#222733",
    "accent":      "#4F8CFF",
    "accent_dim":  "#3B6FD9",
    "text":        "#F5F7FA",
    "text_dim":    "#A7AFBD",
    "ai_bubble":   "#1B2433",
    "user_bubble": "#243B55",
    "border":      "#303746",
    "success":     "#45D483",
    "error":       "#FF5C6C",
    "warning":     "#FFB347",
}

FONT_FAMILY = "Segoe UI"

def f(size=14, bold=False):
    return ctk.CTkFont(family=FONT_FAMILY, size=size, weight="bold" if bold else "normal")

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
EXCELS_DIR  = os.path.normpath(os.path.join(BASE_DIR, "..", "excels"))
WORDS_DIR   = os.path.normpath(os.path.join(BASE_DIR, "..", "words"))
GALLERY_DIR = os.path.normpath(os.path.join(BASE_DIR, "..", "gallery"))
GALLERY_INDEX = os.path.join(GALLERY_DIR, "index.json")


def escanear_biblioteca():
    entries = []
    for carpeta, tipo, ext in [(EXCELS_DIR, "excel", ".xlsx"), (WORDS_DIR, "word", ".docx")]:
        if not os.path.exists(carpeta):
            continue
        for fname in os.listdir(carpeta):
            if fname.endswith(ext):
                fpath = os.path.join(carpeta, fname)
                nombre = fname[:-len(ext)]
                fecha = datetime.fromtimestamp(os.path.getmtime(fpath)).strftime("%Y-%m-%d %H:%M")
                json_path = os.path.join(carpeta, nombre + ".json")
                doc_id = None
                if os.path.exists(json_path):
                    try:
                        with open(json_path, "r", encoding="utf-8") as fh:
                            data_local = json.load(fh)
                        doc_id = data_local.get("_firebase_doc_id")
                    except Exception:
                        pass
                entries.append({
                    "nombre": nombre,
                    "tipo":   tipo,
                    "fecha":  fecha,
                    "path":   fpath,
                    "json_path": json_path,
                    "doc_id": doc_id,
                })
    entries.sort(key=lambda e: e["fecha"], reverse=True)
    return entries


def guardar_json_documento(path_archivo, data):
    base = os.path.splitext(path_archivo)[0]
    with open(base + ".json", "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)


def cargar_json_documento(path_archivo):
    base = os.path.splitext(path_archivo)[0]
    json_path = base + ".json"
    if os.path.exists(json_path):
        with open(json_path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    return None


def cargar_galeria():
    os.makedirs(GALLERY_DIR, exist_ok=True)
    if os.path.exists(GALLERY_INDEX):
        with open(GALLERY_INDEX, "r", encoding="utf-8") as fh:
            return json.load(fh)
    return {}


def guardar_galeria(galeria):
    os.makedirs(GALLERY_DIR, exist_ok=True)
    with open(GALLERY_INDEX, "w", encoding="utf-8") as fh:
        json.dump(galeria, fh, ensure_ascii=False, indent=2)


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("taller.ai")
        self.geometry("900x650")
        self.minsize(760, 560)
        self.configure(fg_color=COLORS["bg"])
        self.current_frame = None
        self.sesion = None

        sesion_guardada = cargar_sesion()
        if sesion_guardada:
            self.sesion = sesion_guardada
            self.mostrar_menu()
        else:
            self.mostrar_login()

    def cambiar_frame(self, frame_class, **kwargs):
        if self.current_frame:
            self.current_frame.destroy()
        self.current_frame = frame_class(self, **kwargs)
        self.current_frame.pack(fill="both", expand=True)

    def mostrar_login(self):
        self.cambiar_frame(LoginFrame)

    def mostrar_menu(self):
        self.cambiar_frame(MenuFrame)

    def mostrar_chat(self, tipo, json_inicial=None, path_inicial=None, prompt_inicial=None, indice_inicial=None):
        self.cambiar_frame(
            ChatFrame,
            tipo=tipo,
            json_inicial=json_inicial,
            path_inicial=path_inicial,
            prompt_inicial=prompt_inicial,
            indice_inicial=indice_inicial
        )

    def mostrar_biblioteca(self):
        self.cambiar_frame(BibliotecaFrame)

    def mostrar_galeria(self):
        self.cambiar_frame(GaleriaFrame)

    def mostrar_empresa(self):
        self.cambiar_frame(EmpresaFrame)

    def mostrar_plantillas(self):
        self.cambiar_frame(PlantillasFrame)


class LoginFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color=COLORS["bg"])
        self.master = master
        self._build_ui()

    def _build_ui(self):
        ctk.CTkLabel(
            self, text="taller.ai",
            font=ctk.CTkFont(family=FONT_FAMILY, size=48, weight="bold"),
            text_color=COLORS["accent"]
        ).pack(pady=(60, 4))

        ctk.CTkLabel(
            self, text="Generador de documentos con IA para tu empresa ficticia",
            font=f(14), text_color=COLORS["text_dim"]
        ).pack(pady=(0, 40))

        card = ctk.CTkFrame(self, fg_color=COLORS["surface"], corner_radius=18, width=420)
        card.pack(padx=40, pady=10)
        card.pack_propagate(False)
        card.configure(height=380)

        ctk.CTkLabel(
            card, text="Acceder a tu grupo",
            font=f(18, True), text_color=COLORS["text"]
        ).pack(pady=(28, 20))

        self.tab_var = ctk.StringVar(value="unirse")
        tabs = ctk.CTkFrame(card, fg_color=COLORS["surface2"], corner_radius=10)
        tabs.pack(padx=24, fill="x")

        self.btn_tab_unirse = ctk.CTkButton(
            tabs, text="Unirse a grupo",
            fg_color=COLORS["accent"], hover_color=COLORS["accent_dim"],
            text_color="#000", font=f(13, True),
            corner_radius=8, height=34,
            command=self._mostrar_unirse
        )
        self.btn_tab_unirse.pack(side="left", expand=True, fill="x", padx=(4, 2), pady=4)

        self.btn_tab_crear = ctk.CTkButton(
            tabs, text="Crear grupo",
            fg_color="transparent", hover_color=COLORS["surface"],
            text_color=COLORS["text_dim"], font=f(13),
            corner_radius=8, height=34,
            command=self._mostrar_crear
        )
        self.btn_tab_crear.pack(side="left", expand=True, fill="x", padx=(2, 4), pady=4)

        self.form_container = ctk.CTkFrame(card, fg_color="transparent")
        self.form_container.pack(fill="both", expand=True, padx=24, pady=12)

        self.error_label = ctk.CTkLabel(
            card, text="", text_color=COLORS["error"], font=f(12)
        )
        self.error_label.pack(pady=(0, 12))

        self._mostrar_unirse()

    def _limpiar_form(self):
        for w in self.form_container.winfo_children():
            w.destroy()
        self.error_label.configure(text="")

    def _mostrar_unirse(self):
        self._limpiar_form()
        self.btn_tab_unirse.configure(fg_color=COLORS["accent"], text_color="#000", font=f(13, True))
        self.btn_tab_crear.configure(fg_color="transparent", text_color=COLORS["text_dim"], font=f(13))

        entry_cfg = dict(
            fg_color=COLORS["surface2"], border_color=COLORS["border"],
            text_color=COLORS["text"], font=f(13), height=40, corner_radius=10
        )

        ctk.CTkLabel(self.form_container, text="Nombre del grupo", font=f(12),
                     text_color=COLORS["text_dim"], anchor="w").pack(fill="x")
        self.entry_nombre_login = ctk.CTkEntry(
            self.form_container, placeholder_text="Ej: Los Piratas", **entry_cfg)
        self.entry_nombre_login.pack(fill="x", pady=(2, 10))

        ctk.CTkLabel(self.form_container, text="Contrasena (opcional)", font=f(12),
                     text_color=COLORS["text_dim"], anchor="w").pack(fill="x")
        self.entry_pass = ctk.CTkEntry(
            self.form_container, placeholder_text="Vacio = grupo sin contrasena",
            show="*", **entry_cfg)
        self.entry_pass.pack(fill="x", pady=(2, 14))

        ctk.CTkButton(
            self.form_container, text="Entrar",
            fg_color=COLORS["accent"], hover_color=COLORS["accent_dim"],
            text_color="#000", font=f(14, True),
            height=42, corner_radius=12,
            command=self._unirse
        ).pack(fill="x")

    def _mostrar_crear(self):
        self._limpiar_form()
        self.btn_tab_crear.configure(fg_color=COLORS["accent"], text_color="#000", font=f(13, True))
        self.btn_tab_unirse.configure(fg_color="transparent", text_color=COLORS["text_dim"], font=f(13))

        entry_cfg = dict(
            fg_color=COLORS["surface2"], border_color=COLORS["border"],
            text_color=COLORS["text"], font=f(13), height=40, corner_radius=10
        )

        ctk.CTkLabel(self.form_container, text="Nombre del grupo", font=f(12),
                     text_color=COLORS["text_dim"], anchor="w").pack(fill="x")
        self.entry_nombre = ctk.CTkEntry(
            self.form_container, placeholder_text="Ej: Los Piratas S.A.", **entry_cfg)
        self.entry_nombre.pack(fill="x", pady=(2, 10))

        ctk.CTkLabel(self.form_container, text="Contrasena (opcional)", font=f(12),
                     text_color=COLORS["text_dim"], anchor="w").pack(fill="x")
        self.entry_pass_crear = ctk.CTkEntry(
            self.form_container, placeholder_text="Vacio = grupo publico, cualquiera entra",
            show="*", **entry_cfg)
        self.entry_pass_crear.pack(fill="x", pady=(2, 14))

        ctk.CTkButton(
            self.form_container, text="Crear grupo",
            fg_color=COLORS["success"], hover_color="#35B563",
            text_color="#000", font=f(14, True),
            height=42, corner_radius=12,
            command=self._crear
        ).pack(fill="x")

    def _unirse(self):
        nombre  = self.entry_nombre_login.get().strip()
        password = self.entry_pass.get().strip()

        if not nombre:
            self.error_label.configure(text="Ingresa el nombre del grupo")
            return

        self.error_label.configure(text="Conectando...", text_color=COLORS["text_dim"])
        self.update()

        def _worker():
            from back.firebase import unirse_grupo
            resultado = unirse_grupo(nombre, password)
            self.after(0, lambda: self._on_unirse(resultado))

        threading.Thread(target=_worker, daemon=True).start()

    def _on_unirse(self, resultado):
        if resultado is None:
            self.error_label.configure(text="Grupo no encontrado o contrasena incorrecta", text_color=COLORS["error"])
            return
        self.master.sesion = resultado
        guardar_sesion(resultado["codigo"], resultado["nombre_grupo"], resultado.get("empresa", {}))
        self.master.mostrar_menu()

    def _crear(self):
        nombre   = self.entry_nombre.get().strip()
        password = self.entry_pass_crear.get().strip()

        if not nombre:
            self.error_label.configure(text="Ingresa el nombre del grupo")
            return

        self.error_label.configure(text="Creando grupo...", text_color=COLORS["text_dim"])
        self.update()

        def _worker():
            from back.firebase import crear_grupo
            resultado = crear_grupo(nombre, password)
            self.after(0, lambda: self._on_crear(resultado, nombre))

        threading.Thread(target=_worker, daemon=True).start()

    def _on_crear(self, resultado, nombre):
        if resultado is None:
            self.error_label.configure(
                text="Ya existe un grupo con ese nombre, o sin conexion.", text_color=COLORS["error"])
            return

        sesion = {"codigo": nombre, "nombre_grupo": nombre, "empresa": {}}
        self.master.sesion = sesion
        guardar_sesion(nombre, nombre, {})

        self._limpiar_form()
        self.error_label.configure(text="")

        ctk.CTkLabel(
            self.form_container,
            text="Grupo creado exitosamente",
            font=f(14, True), text_color=COLORS["success"]
        ).pack(pady=(8, 4))

        ctk.CTkLabel(
            self.form_container,
            text=f"Tu grupo '{nombre}' ya esta listo.\nCualquiera puede unirse con este nombre.",
            font=f(12), text_color=COLORS["text_dim"]
        ).pack(pady=(0, 8))

        ctk.CTkButton(
            self.form_container, text="Continuar",
            fg_color=COLORS["accent"], hover_color=COLORS["accent_dim"],
            text_color="#000", font=f(14, True),
            height=42, corner_radius=12,
            command=self.master.mostrar_menu
        ).pack(fill="x", pady=(10, 0))


class MenuFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color=COLORS["bg"])
        self.master = master

        sesion = master.sesion or {}
        nombre_grupo = sesion.get("nombre_grupo", "")
        empresa = sesion.get("empresa", {})
        tiene_empresa = empresa_completa(empresa)

        header = ctk.CTkFrame(self, fg_color=COLORS["surface"], height=56, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkLabel(
            header,
            text=f"Grupo: {nombre_grupo}" if nombre_grupo else "taller.ai",
            font=f(13, True), text_color=COLORS["text"]
        ).pack(side="left", padx=16, pady=8)

        ctk.CTkButton(
            header, text="Cerrar sesion",
            width=110, height=30,
            fg_color="transparent", hover_color=COLORS["error"],
            text_color=COLORS["text_dim"], font=f(11),
            corner_radius=8,
            command=self._cerrar_sesion
        ).pack(side="right", padx=12, pady=8)

        ctk.CTkLabel(
            self, text="taller.ai",
            font=ctk.CTkFont(family=FONT_FAMILY, size=42, weight="bold"),
            text_color=COLORS["accent"]
        ).pack(pady=(40, 4))

        ctk.CTkLabel(
            self, text="Generador de documentos con IA",
            font=f(14), text_color=COLORS["text_dim"]
        ).pack(pady=(0, 30))

        btn_cfg = dict(width=280, height=52, font=f(15, True), corner_radius=14)

        ctk.CTkButton(
            self, text="Plantillas rapidas",
            fg_color=COLORS["accent"], hover_color=COLORS["accent_dim"],
            text_color="#000",
            command=master.mostrar_plantillas,
            **btn_cfg
        ).pack(pady=6)

        row2 = ctk.CTkFrame(self, fg_color="transparent")
        row2.pack(pady=4)

        ctk.CTkButton(
            row2, text="Generar Excel",
            fg_color=COLORS["surface2"], hover_color=COLORS["accent"],
            text_color=COLORS["text"],
            command=lambda: master.mostrar_chat("excel"),
            width=134, height=46, font=f(13, True), corner_radius=14
        ).pack(side="left", padx=4)

        ctk.CTkButton(
            row2, text="Generar Word",
            fg_color=COLORS["surface2"], hover_color=COLORS["accent"],
            text_color=COLORS["text"],
            command=lambda: master.mostrar_chat("word"),
            width=134, height=46, font=f(13, True), corner_radius=14
        ).pack(side="left", padx=4)

        row3 = ctk.CTkFrame(self, fg_color="transparent")
        row3.pack(pady=4)

        ctk.CTkButton(
            row3, text="Biblioteca",
            fg_color=COLORS["surface"], hover_color=COLORS["surface2"],
            text_color=COLORS["text_dim"],
            command=master.mostrar_biblioteca,
            width=88, height=40, font=f(12), corner_radius=12
        ).pack(side="left", padx=4)

        ctk.CTkButton(
            row3, text="Galeria",
            fg_color=COLORS["surface"], hover_color=COLORS["surface2"],
            text_color=COLORS["text_dim"],
            command=master.mostrar_galeria,
            width=88, height=40, font=f(12), corner_radius=12
        ).pack(side="left", padx=4)

        emp_color  = COLORS["surface"] if tiene_empresa else COLORS["warning"]
        emp_tcolor = COLORS["text_dim"] if tiene_empresa else "#000"
        emp_texto  = "Mi Empresa" if tiene_empresa else "Mi Empresa !"

        ctk.CTkButton(
            row3, text=emp_texto,
            fg_color=emp_color, hover_color=COLORS["surface2"],
            text_color=emp_tcolor,
            command=master.mostrar_empresa,
            width=100, height=40, font=f(12, not tiene_empresa), corner_radius=12
        ).pack(side="left", padx=4)

        if not tiene_empresa:
            ctk.CTkLabel(
                self,
                text="Configura tu empresa para que aparezca en todos los documentos",
                font=f(11), text_color=COLORS["warning"]
            ).pack(pady=(6, 0))

    def _cerrar_sesion(self):
        cerrar_sesion()
        self.master.sesion = None
        self.master.mostrar_login()


class EmpresaFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color=COLORS["bg"])
        self.master = master
        self.entries = {}
        self._build_ui()

    def _build_ui(self):
        header = ctk.CTkFrame(self, fg_color=COLORS["surface"], height=72, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkButton(
            header, text="←", width=40, height=36,
            fg_color="transparent", hover_color=COLORS["surface2"],
            text_color=COLORS["text_dim"],
            command=self.master.mostrar_menu, font=f(18, True)
        ).pack(side="left", padx=8, pady=8)

        ctk.CTkLabel(
            header, text="Datos de la Empresa",
            font=f(15, True), text_color=COLORS["text"]
        ).pack(side="left", padx=4)

        scroll = ctk.CTkScrollableFrame(self, fg_color=COLORS["bg"])
        scroll.pack(fill="both", expand=True, padx=60, pady=20)

        empresa_actual = (self.master.sesion or {}).get("empresa", {})

        entry_cfg = dict(
            fg_color=COLORS["surface2"], border_color=COLORS["border"],
            text_color=COLORS["text"], font=f(13), height=42, corner_radius=10
        )

        for campo, label in CAMPOS_EMPRESA:
            ctk.CTkLabel(scroll, text=label, font=f(12),
                         text_color=COLORS["text_dim"], anchor="w").pack(fill="x", pady=(10, 2))
            entry = ctk.CTkEntry(scroll, placeholder_text=label, **entry_cfg)
            entry.pack(fill="x")
            valor_actual = empresa_actual.get(campo, "")
            if valor_actual:
                entry.insert(0, valor_actual)
            self.entries[campo] = entry

        self.status_label = ctk.CTkLabel(scroll, text="", font=f(12))
        self.status_label.pack(pady=(16, 4))

        ctk.CTkButton(
            scroll, text="Guardar",
            fg_color=COLORS["accent"], hover_color=COLORS["accent_dim"],
            text_color="#000", font=f(14, True),
            height=46, corner_radius=12,
            command=self._guardar
        ).pack(fill="x", pady=(4, 20))

    def _guardar(self):
        empresa = {campo: self.entries[campo].get().strip() for campo, _ in CAMPOS_EMPRESA}

        self.status_label.configure(text="Guardando...", text_color=COLORS["text_dim"])
        self.update()

        def _worker():
            from back.firebase import guardar_empresa
            sesion = self.master.sesion or {}
            codigo = sesion.get("codigo", "")
            ok = True
            if codigo:
                ok = guardar_empresa(codigo, empresa)
            actualizar_empresa_sesion(empresa)
            if self.master.sesion:
                self.master.sesion["empresa"] = empresa
            self.after(0, lambda: self._on_guardado(ok))

        threading.Thread(target=_worker, daemon=True).start()

    def _on_guardado(self, ok):
        if ok:
            self.status_label.configure(
                text="Empresa guardada correctamente", text_color=COLORS["success"])
        else:
            self.status_label.configure(
                text="Guardado localmente (sin conexion a Firebase)", text_color=COLORS["warning"])


class PlantillasFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color=COLORS["bg"])
        self.master = master
        self._build_ui()

    def _build_ui(self):
        header = ctk.CTkFrame(self, fg_color=COLORS["surface"], height=72, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkButton(
            header, text="←", width=40, height=36,
            fg_color="transparent", hover_color=COLORS["surface2"],
            text_color=COLORS["text_dim"],
            command=self.master.mostrar_menu, font=f(18, True)
        ).pack(side="left", padx=8, pady=8)

        ctk.CTkLabel(
            header, text="Plantillas Rapidas",
            font=f(15, True), text_color=COLORS["text"]
        ).pack(side="left", padx=4)

        ctk.CTkLabel(
            self,
            text="Genera documentos completos con un solo click",
            font=f(13), text_color=COLORS["text_dim"]
        ).pack(pady=(16, 8))

        scroll = ctk.CTkScrollableFrame(self, fg_color=COLORS["bg"])
        scroll.pack(fill="both", expand=True, padx=24, pady=8)

        plantillas = [
            {
                "icon":  "🧾",
                "titulo": "Factura de venta",
                "desc":  "Excel con items, IVA 21%, subtotal y total",
                "tipo":  "excel",
                "action": self._dialogo_factura,
            },
            {
                "icon":  "📦",
                "titulo": "Lista de productos",
                "desc":  "Catalogo con codigo, descripcion, precio y stock",
                "tipo":  "excel",
                "action": self._dialogo_productos,
            },
            {
                "icon":  "🚚",
                "titulo": "Proveedores",
                "desc":  "Planilla con contactos, productos y condiciones de pago",
                "tipo":  "excel",
                "action": self._dialogo_proveedores,
            },
            {
                "icon":  "👥",
                "titulo": "Base de clientes",
                "desc":  "Listado de clientes con datos de contacto y categoria",
                "tipo":  "excel",
                "action": self._dialogo_clientes,
            },
            {
                "icon":  "📈",
                "titulo": "Planilla de ventas",
                "desc":  "Registro del periodo con totales y grafico",
                "tipo":  "excel",
                "action": self._dialogo_ventas,
            },
            {
                "icon":  "📄",
                "titulo": "Informe Word",
                "desc":  "Documento formal con portada, indice y secciones",
                "tipo":  "word",
                "action": self._dialogo_informe,
            },
            {
                "icon":  "📋",
                "titulo": "Catalogo Word",
                "desc":  "Lista de productos/servicios con descripciones detalladas",
                "tipo":  "word",
                "action": self._dialogo_catalogo,
            },
        ]

        for p in plantillas:
            self._agregar_tarjeta(scroll, p)

    def _agregar_tarjeta(self, parent, p):
        card = ctk.CTkFrame(parent, fg_color=COLORS["surface"], corner_radius=14)
        card.pack(fill="x", pady=6)

        left = ctk.CTkFrame(card, fg_color="transparent")
        left.pack(side="left", fill="both", expand=True, padx=16, pady=14)

        ctk.CTkLabel(
            left, text=f"{p['icon']}  {p['titulo']}",
            font=f(14, True), text_color=COLORS["text"], anchor="w"
        ).pack(anchor="w")

        tipo_color = COLORS["success"] if p["tipo"] == "excel" else COLORS["accent"]
        tipo_texto = "Excel" if p["tipo"] == "excel" else "Word"

        row = ctk.CTkFrame(left, fg_color="transparent")
        row.pack(anchor="w", pady=(2, 0))

        ctk.CTkLabel(
            row, text=tipo_texto,
            font=f(10, True), text_color=tipo_color
        ).pack(side="left")

        ctk.CTkLabel(
            row, text=f"  {p['desc']}",
            font=f(11), text_color=COLORS["text_dim"]
        ).pack(side="left")

        ctk.CTkButton(
            card, text="Generar",
            width=90, height=36,
            fg_color=COLORS["accent"], hover_color=COLORS["accent_dim"],
            text_color="#000", font=f(12, True), corner_radius=10,
            command=p["action"]
        ).pack(side="right", padx=16)

    def _abrir_dialogo(self, titulo, campos, callback):
        top = ctk.CTkToplevel(self)
        top.title(titulo)
        top.geometry("420x" + str(100 + len(campos) * 72 + 160))
        top.configure(fg_color=COLORS["bg"])
        top.grab_set()

        ctk.CTkLabel(top, text=titulo, font=f(15, True),
                     text_color=COLORS["text"]).pack(pady=(20, 4))

        ctk.CTkLabel(top, text="Todos los campos son opcionales.",
                     font=f(11), text_color=COLORS["text_dim"]).pack(pady=(0, 12))

        entries = {}
        entry_cfg = dict(
            fg_color=COLORS["surface2"], border_color=COLORS["border"],
            text_color=COLORS["text"], font=f(13), height=40, corner_radius=10
        )

        form = ctk.CTkFrame(top, fg_color="transparent")
        form.pack(fill="both", expand=True, padx=24)

        for key, label, placeholder in campos:
            ctk.CTkLabel(form, text=f"{label} (opcional)", font=f(11),
                         text_color=COLORS["text_dim"], anchor="w").pack(fill="x")
            e = ctk.CTkEntry(form, placeholder_text=placeholder, **entry_cfg)
            e.pack(fill="x", pady=(2, 8))
            entries[key] = e

        ctk.CTkLabel(form, text="Anotación / notas adicionales (opcional)",
                     font=f(11), text_color=COLORS["text_dim"], anchor="w").pack(fill="x", pady=(8, 0))
        nota_box = ctk.CTkTextbox(
            form, height=80,
            fg_color=COLORS["surface2"], border_color=COLORS["border"], border_width=1,
            text_color=COLORS["text"], font=f(12), corner_radius=10
        )
        nota_box.pack(fill="x", pady=(2, 0))

        def confirmar():
            valores = {k: v.get().strip() for k, v in entries.items()}
            nota = nota_box.get("1.0", "end").strip()
            top.destroy()
            callback({**valores, "nota": nota})

        ctk.CTkButton(
            top, text="Generar documento",
            fg_color=COLORS["accent"], hover_color=COLORS["accent_dim"],
            text_color="#000", font=f(13, True),
            height=42, corner_radius=12,
            command=confirmar
        ).pack(padx=24, fill="x", pady=12)

    def _abrir_dialogo_multi(self, titulo, layout_filas, campos_globales, callback):
        top = ctk.CTkToplevel(self)
        top.title(titulo)
        top.geometry("900x650")
        top.configure(fg_color=COLORS["bg"])
        top.grab_set()

        ctk.CTkLabel(top, text=titulo, font=f(15, True),
                     text_color=COLORS["text"]).pack(pady=(20, 4))

        ctk.CTkLabel(top, text="Tocá '+ Agregar' para sumar filas. Todos los campos son opcionales.",
                     font=f(11), text_color=COLORS["text_dim"]).pack(pady=(0, 10))

        global_entries = {}
        entry_cfg = dict(
            fg_color=COLORS["surface2"], border_color=COLORS["border"],
            text_color=COLORS["text"], font=f(12), height=36, corner_radius=8
        )

        if campos_globales:
            global_frame = ctk.CTkFrame(top, fg_color=COLORS["surface"], corner_radius=10)
            global_frame.pack(fill="x", padx=24, pady=(0, 10))

            for key, label, placeholder in campos_globales:
                ctk.CTkLabel(global_frame, text=f"{label} (opcional)", font=f(11),
                             text_color=COLORS["text_dim"], anchor="w").pack(fill="x", padx=12, pady=(8, 0))
                e = ctk.CTkEntry(global_frame, placeholder_text=placeholder, **entry_cfg)
                e.pack(fill="x", padx=12, pady=(2, 6))
                global_entries[key] = e

        if layout_filas:
            header = ctk.CTkFrame(top, fg_color="transparent")
            header.pack(fill="x", padx=24)
            ncols_hdr = len(layout_filas[0]) + 1
            for i in range(ncols_hdr):
                header.grid_columnconfigure(i, weight=1, uniform="hdr")
            for i, (_, lbl, _) in enumerate(layout_filas[0]):
                ctk.CTkLabel(header, text=lbl, font=f(10, True), text_color=COLORS["text_dim"],
                             anchor="w").grid(row=0, column=i, sticky="w", padx=8)

        scroll = ctk.CTkScrollableFrame(top, fg_color=COLORS["bg"])
        scroll.pack(fill="both", expand=True, padx=24, pady=8)

        filas_data = []
        nrows_layout = len(layout_filas)
        ncols_top = len(layout_filas[0])

        def agregar_fila():
            item = ctk.CTkFrame(scroll, fg_color=COLORS["surface"], corner_radius=10)
            item.pack(fill="x", pady=4)

            for i in range(ncols_top + 1):
                item.grid_columnconfigure(i, weight=1, uniform="it")

            all_entries = {}

            for i, (key, label, placeholder) in enumerate(layout_filas[0]):
                e = ctk.CTkEntry(item, placeholder_text=placeholder, **entry_cfg)
                e.grid(row=0, column=i, sticky="ew", padx=(8, 4), pady=(8, 4))
                all_entries[key] = e

            def eliminar():
                filas_data.remove(fila)
                item.destroy()

            ctk.CTkButton(
                item, text="X", width=32, height=32,
                fg_color=COLORS["surface2"], hover_color=COLORS["error"],
                text_color=COLORS["text_dim"], font=f(11, True), corner_radius=8,
                command=eliminar
            ).grid(row=0, column=ncols_top, padx=(4, 8), pady=(8, 4), rowspan=nrows_layout)

            for r, fila_campos in enumerate(layout_filas[1:], start=1):
                for i, (key, label, placeholder) in enumerate(fila_campos):
                    e = ctk.CTkEntry(item, placeholder_text=placeholder, **entry_cfg)
                    e.grid(row=r, column=i, sticky="ew", padx=(8, 4), pady=(0, 8))
                    all_entries[key] = e

            fila = {"frame": item, "entries": all_entries}
            filas_data.append(fila)

        btn_frame = ctk.CTkFrame(top, fg_color="transparent")
        btn_frame.pack(fill="x", padx=24, pady=4)

        ctk.CTkButton(
            btn_frame, text="+ Agregar",
            fg_color=COLORS["surface2"], hover_color=COLORS["accent"],
            text_color=COLORS["text"], font=f(12, True), corner_radius=10,
            command=agregar_fila
        ).pack(side="left")

        nota_frame = ctk.CTkFrame(top, fg_color="transparent")
        nota_frame.pack(fill="x", padx=24, pady=(8, 4))

        ctk.CTkLabel(nota_frame, text="Anotación / notas adicionales (opcional)",
                     font=f(11), text_color=COLORS["text_dim"], anchor="w").pack(fill="x")
        nota_box = ctk.CTkTextbox(
            nota_frame, height=60,
            fg_color=COLORS["surface2"], border_color=COLORS["border"], border_width=1,
            text_color=COLORS["text"], font=f(12), corner_radius=10
        )
        nota_box.pack(fill="x", pady=(2, 0))

        def confirmar():
            globals_data = {k: v.get().strip() for k, v in global_entries.items()}
            items = []
            for fila in filas_data:
                item_data = {k: v.get().strip() for k, v in fila["entries"].items()}
                if any(item_data.values()):
                    items.append(item_data)
            nota = nota_box.get("1.0", "end").strip()
            top.destroy()
            callback({**globals_data, "items": items, "nota": nota})

        ctk.CTkButton(
            top, text="Generar documento",
            fg_color=COLORS["accent"], hover_color=COLORS["accent_dim"],
            text_color="#000", font=f(13, True),
            height=42, corner_radius=12,
            command=confirmar
        ).pack(padx=24, fill="x", pady=12)

        agregar_fila()

    def _dialogo_factura(self):
        self._abrir_dialogo(
            "Nueva Factura",
            [
                ("cliente",       "Cliente",                "Nombre del cliente o empresa"),
                ("items",         "Productos / servicios",  "Ej: 3 sillas $5000, 1 mesa $12000"),
                ("condicion",     "Condicion de pago",      "Ej: contado, 30 dias, cuotas"),
            ],
            lambda v: self._lanzar(
                "excel",
                __import__("back.plantillas", fromlist=["prompt_factura"]).prompt_factura(
                    cliente=v.get("cliente", ""),
                    items=v.get("items", ""),
                    condicion_pago=v.get("condicion", ""),
                    sesion=self.master.sesion,
                    nota=v.get("nota", ""),
                )
            )
        )

    def _dialogo_productos(self):
        self._abrir_dialogo_multi(
            "Lista de Productos",
            [
                [
                    ("nombre",      "Nombre",           "Ej: taragui 500g"),
                    ("descripcion", "Descripción",      "—"),
                    ("precio",      "Precio Unitario",  "Ej: 1500"),
                    ("stock",       "Stock (cantidad)", "Ej: 20"),
                ]
            ],
            [("categoria", "Categoría de productos", "Ej: electrodomésticos, indumentaria")],
            lambda v: self._lanzar(
                "excel",
                __import__("back.plantillas", fromlist=["prompt_productos"]).prompt_productos(
                    productos=v.get("items", []),
                    categoria=v.get("categoria", ""),
                    sesion=self.master.sesion,
                    nota=v.get("nota", ""),
                )
            )
        )

    def _dialogo_proveedores(self):
        self._abrir_dialogo_multi(
            "Planilla de Proveedores",
            [
                [
                    ("nombre",     "Nombre",     "—"),
                    ("cuit",       "CUIT",       "—"),
                    ("contacto",   "Contacto",   "—"),
                    ("telefono",   "Teléfono",   "—"),
                    ("email",      "Email",      "—"),
                ],
                [
                    ("productos",   "Productos que provee", "—"),
                    ("condiciones", "Condiciones de pago",  "—"),
                    ("tiempo",      "Tiempo de entrega",    "—"),
                ]
            ],
            [],
            lambda v: self._lanzar(
                "excel",
                __import__("back.plantillas", fromlist=["prompt_proveedores"]).prompt_proveedores(
                    proveedores=v.get("items", []),
                    sesion=self.master.sesion,
                    nota=v.get("nota", ""),
                )
            )
        )

    def _dialogo_clientes(self):
        self._abrir_dialogo_multi(
            "Base de Clientes",
            [
                [
                    ("nombre",    "Nombre",       "—"),
                    ("direccion", "Dirección",    "—"),
                    ("dni",       "DNI/CUIT",     "—"),
                    ("telefono",  "Teléfono",     "—"),
                    ("email",     "Email",        "—"),
                ],
                [
                    ("ciudad",    "Ciudad",       "—"),
                    ("categoria", "Categoría",    "—"),
                    ("contenido", "Contenido",    "—"),
                ]
            ],
            [],
            lambda v: self._lanzar(
                "excel",
                __import__("back.plantillas", fromlist=["prompt_clientes"]).prompt_clientes(
                    clientes=v.get("items", []),
                    sesion=self.master.sesion,
                    nota=v.get("nota", ""),
                )
            )
        )

    def _dialogo_ventas(self):
        self._abrir_dialogo_multi(
            "Planilla de Ventas",
            [
                [
                    ("cliente",   "Cliente",  "—"),
                    ("producto",  "Producto", "—"),
                    ("vendedor",  "Vendedor", "—"),
                ],
                [
                    ("cantidad", "Cantidad", "—"),
                    ("precio",   "Precio",   "—"),
                    ("fecha",    "Fecha (DD/MM/YYYY)", "—"),
                ]
            ],
            [("periodo", "Período", "Ej: enero 2025, primer trimestre 2025")],
            lambda v: self._lanzar(
                "excel",
                __import__("back.plantillas", fromlist=["prompt_ventas"]).prompt_ventas(
                    ventas=v.get("items", []),
                    periodo=v.get("periodo", ""),
                    sesion=self.master.sesion,
                    nota=v.get("nota", ""),
                )
            )
        )

    def _dialogo_informe(self):
        self._abrir_dialogo_multi(
            "Informe Word",
            [[("seccion", "Sección", "Ej: Introducción")]],
            [("tema", "Tema del informe", "Ej: análisis de mercado, plan de negocios")],
            lambda v: self._lanzar_con_indice(
                "word",
                *__import__("back.plantillas", fromlist=["prompt_informe_word"]).prompt_informe_word(
                    tema=v.get("tema", ""),
                    secciones=[item["seccion"] for item in v.get("items", [])],
                    sesion=self.master.sesion,
                    nota=v.get("nota", ""),
                )
            )
        )

    def _dialogo_catalogo(self):
        self._abrir_dialogo_multi(
            "Catalogo Word",
            [
                [
                    ("nombre",      "Nombre",         "—"),
                    ("descripcion", "Descripción",    "—"),
                    ("palabras",    "Palabras clave", "—"),
                ]
            ],
            [],
            lambda v: self._lanzar(
                "word",
                __import__("back.plantillas", fromlist=["prompt_catalogo_word"]).prompt_catalogo_word(
                    productos=v.get("items", []),
                    sesion=self.master.sesion,
                    nota=v.get("nota", ""),
                )
            )
        )

    def _lanzar(self, tipo, prompt):
        self.master.mostrar_chat(tipo=tipo, prompt_inicial=prompt)

    def _lanzar_con_indice(self, tipo, prompt, indice):
        self.master.mostrar_chat(tipo=tipo, prompt_inicial=prompt, indice_inicial=indice)


class ChatFrame(ctk.CTkFrame):
    def __init__(self, master, tipo, json_inicial=None, path_inicial=None, prompt_inicial=None, indice_inicial=None):
        super().__init__(master, fg_color=COLORS["bg"])
        self.master = master
        self.tipo = tipo
        self.seleccion = "1" if tipo == "excel" else "2"
        self.ultimo_json = json_inicial
        self.ultimo_path = path_inicial
        self.prompt_inicial = prompt_inicial
        self.indice_pendiente = indice_inicial
        self.historial_mensajes = []

        if json_inicial and path_inicial:
            self.estado = "editando"
        else:
            self.estado = "esperando_prompt"

        self._build_ui()
        self.after(300, self._saludo_inicial)

    def _build_ui(self):
        header = ctk.CTkFrame(self, fg_color=COLORS["surface"], height=72, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkButton(
            header, text="←", width=40, height=36,
            fg_color="transparent", hover_color=COLORS["surface2"],
            text_color=COLORS["text_dim"],
            command=self.master.mostrar_menu, font=f(18, True)
        ).pack(side="left", padx=8, pady=8)

        tipo_str = "Excel" if self.tipo == "excel" else "Word"
        modo = " — Editando" if self.estado == "editando" else ""
        ctk.CTkLabel(
            header, text=f"Generar {tipo_str}{modo}",
            font=f(15, True), text_color=COLORS["text"]
        ).pack(side="left", padx=4)

        sesion = self.master.sesion or {}
        if sesion.get("nombre_grupo"):
            ctk.CTkLabel(
                header, text=sesion["nombre_grupo"],
                font=f(11), text_color=COLORS["text_dim"]
            ).pack(side="right", padx=16)

        self.chat_scroll = ctk.CTkScrollableFrame(self, fg_color=COLORS["bg"], corner_radius=0)
        self.chat_scroll.pack(fill="both", expand=True)

        input_bar = ctk.CTkFrame(self, fg_color=COLORS["surface"], height=72, corner_radius=0)
        input_bar.pack(fill="x")
        input_bar.pack_propagate(False)

        self.input_box = ctk.CTkEntry(
            input_bar,
            placeholder_text="Describe tu documento... (usa /imagen para insertar imagenes)",
            fg_color=COLORS["surface2"], border_color=COLORS["border"],
            text_color=COLORS["text"], font=f(13),
            height=40, corner_radius=14
        )
        self.input_box.pack(side="left", fill="x", expand=True, padx=(12, 8), pady=12)
        self.input_box.bind("<Return>", lambda e: self._manejar_enter())
        self.input_box.bind("<KeyRelease>", lambda e: self._actualizar_autocompletado())
        self.input_box.bind("<Up>",   lambda e: self._navegar_sugerencias(-1))
        self.input_box.bind("<Down>", lambda e: self._navegar_sugerencias(1))

        self.sugerencias_frame   = None
        self.sugerencias_lista   = []
        self.indice_seleccion    = -1
        self.sugerencias_botones = []
        self.ultima_barra        = -1

        self.btn_enviar = ctk.CTkButton(
            input_bar, text="Enviar",
            width=90, height=40,
            fg_color=COLORS["accent"], hover_color=COLORS["accent_dim"],
            text_color="#000000",
            font=ctk.CTkFont(size=13, weight="bold"),
            corner_radius=14, command=self._enviar
        )
        self.btn_enviar.pack(side="right", padx=(0, 12), pady=12)

    def _agregar_burbuja(self, texto, es_ia=True):
        color  = COLORS["ai_bubble"] if es_ia else COLORS["user_bubble"]
        anchor = "w" if es_ia else "e"
        prefix = "  " if es_ia else ""

        wrapper = ctk.CTkFrame(self.chat_scroll, fg_color="transparent")
        wrapper.pack(fill="x", pady=4, padx=12)

        ctk.CTkLabel(
            wrapper, text=prefix + texto,
            fg_color=color, text_color=COLORS["text"], font=f(13),
            corner_radius=10, wraplength=700, justify="left",
            anchor="w", padx=14, pady=10
        ).pack(anchor=anchor)
        self.after(50, lambda: self.chat_scroll._parent_canvas.yview_moveto(1.0))

    def _agregar_boton_abrir(self, path):
        wrapper = ctk.CTkFrame(self.chat_scroll, fg_color="transparent")
        wrapper.pack(fill="x", pady=4, padx=12)
        ctk.CTkButton(
            wrapper, text="Abrir archivo",
            width=160, height=36,
            fg_color=COLORS["accent"], hover_color=COLORS["accent_dim"],
            text_color="#000000",
            font=ctk.CTkFont(size=13, weight="bold"),
            corner_radius=14,
            command=lambda: os.startfile(path) if os.path.exists(path) else None
        ).pack(anchor="w")
        self.after(50, lambda: self.chat_scroll._parent_canvas.yview_moveto(1.0))

    def _set_input(self, habilitado):
        state = "normal" if habilitado else "disabled"
        self.input_box.configure(state=state)
        self.btn_enviar.configure(state=state)

    def _extraer_imagenes(self, prompt):
        galeria = cargar_galeria()
        matches = re.findall(r'/(\w+)(?::(\d+))?', prompt)
        imagenes = []
        prompt_limpio = prompt

        for match in matches:
            nombre_img, escala_str = match
            if nombre_img in galeria:
                try:
                    escala = int(escala_str) if escala_str else 12
                    escala = max(5, min(18, escala))
                except ValueError:
                    escala = 12

                imagenes.append({
                    "nombre": nombre_img,
                    "path":   galeria[nombre_img]["path"],
                    "escala": escala
                })
                if escala_str:
                    prompt_limpio = re.sub(rf'/\b{nombre_img}\b:\d+', '', prompt_limpio)
                else:
                    prompt_limpio = re.sub(rf'/\b{nombre_img}\b(?::\d+)?', '', prompt_limpio)

        return imagenes, prompt_limpio.strip()

    def _construir_contexto_historial(self):
        if not self.historial_mensajes or len(self.historial_mensajes) <= 1:
            return ""
        mensajes_previos = self.historial_mensajes[:-1]
        if not mensajes_previos:
            return ""
        contexto = "\nCONTEXT FROM PREVIOUS MESSAGES:\n"
        for i, msg in enumerate(mensajes_previos, 1):
            contexto += f"{i}. {msg}\n"
        contexto += "---\n"
        return contexto

    def _actualizar_autocompletado(self):
        texto = self.input_box.get()
        self.ultima_barra = texto.rfind('/')
        if self.ultima_barra == -1:
            self._ocultar_sugerencias()
            return

        pos_despues = self.ultima_barra + 1
        if pos_despues > len(texto):
            self._ocultar_sugerencias()
            return

        partes  = texto[pos_despues:].split()
        prefijo = partes[0] if partes else ""

        galeria = cargar_galeria()
        if not galeria:
            self._ocultar_sugerencias()
            return

        self.sugerencias_lista = [
            img for img in galeria.keys()
            if img.lower().startswith(prefijo.lower())
        ]

        if self.sugerencias_lista:
            self._mostrar_sugerencias(self.sugerencias_lista)
            self.indice_seleccion = 0
        else:
            self._ocultar_sugerencias()

    def _mostrar_sugerencias(self, sugerencias):
        self._ocultar_sugerencias()
        self.input_box.update_idletasks()
        try:
            x = self.input_box.winfo_rootx() - self.winfo_rootx()
            y = self.input_box.winfo_rooty() - self.winfo_rooty()
            w = self.input_box.winfo_width()
            h = min(len(sugerencias), 5) * 32 + 12
            y -= h + 8
        except Exception:
            return

        self.sugerencias_frame = ctk.CTkFrame(
            self, fg_color=COLORS["surface"],
            border_color=COLORS["accent"], border_width=1,
            corner_radius=12, width=w, height=h
        )
        self.sugerencias_frame.place(x=x, y=max(0, y))
        self.sugerencias_frame.lift()

        scroll_frame = ctk.CTkFrame(self.sugerencias_frame, fg_color=COLORS["surface"])
        scroll_frame.pack(fill="both", expand=True, padx=6, pady=6)

        self.sugerencias_botones = []
        for i, sugerencia in enumerate(sugerencias[:8]):
            btn = ctk.CTkButton(
                scroll_frame, text=f"  /{sugerencia}",
                text_color=COLORS["accent"] if i == self.indice_seleccion else COLORS["text"],
                fg_color=COLORS["surface2"] if i == self.indice_seleccion else COLORS["bg"],
                hover_color=COLORS["accent_dim"], font=f(11),
                height=28, corner_radius=4,
                command=lambda s=sugerencia: self._insertar_sugerencia(s)
            )
            btn.pack(fill="x", padx=0, pady=2)
            btn.indice = i
            self.sugerencias_botones.append(btn)

    def _ocultar_sugerencias(self):
        if self.sugerencias_frame:
            try:
                self.sugerencias_frame.destroy()
            except Exception:
                pass
            self.sugerencias_frame   = None
            self.sugerencias_botones = []
        self.indice_seleccion = -1

    def _manejar_enter(self):
        if (self.sugerencias_frame and
                0 <= self.indice_seleccion < len(self.sugerencias_lista)):
            self._insertar_sugerencia(self.sugerencias_lista[self.indice_seleccion])
            return "break"
        self._enviar()

    def _navegar_sugerencias(self, direccion):
        if not self.sugerencias_lista or not self.sugerencias_frame:
            return "break"
        if 0 <= self.indice_seleccion < len(self.sugerencias_botones):
            self.sugerencias_botones[self.indice_seleccion].configure(
                fg_color=COLORS["bg"], text_color=COLORS["text"])

        self.indice_seleccion = max(0, min(
            self.indice_seleccion + direccion, len(self.sugerencias_lista) - 1))

        if 0 <= self.indice_seleccion < len(self.sugerencias_botones):
            self.sugerencias_botones[self.indice_seleccion].configure(
                fg_color=COLORS["surface2"], text_color=COLORS["accent"])
        return "break"

    def _insertar_sugerencia(self, sugerencia):
        texto_actual = self.input_box.get()
        pos_barra    = self.ultima_barra
        pos_texto    = pos_barra + 1

        escala = ""
        if pos_texto < len(texto_actual):
            resto = texto_actual[pos_texto:].split()[0] if texto_actual[pos_texto:].split() else ""
            if ':' in resto:
                escala = ":" + resto.split(':')[1]

        texto_nuevo = texto_actual[:pos_barra] + "/" + sugerencia + escala

        if pos_texto < len(texto_actual):
            fin_palabra = pos_texto
            while fin_palabra < len(texto_actual) and texto_actual[fin_palabra] not in ' \t':
                fin_palabra += 1
            texto_nuevo += texto_actual[fin_palabra:]

        self.input_box.delete(0, "end")
        self.input_box.insert(0, texto_nuevo)
        self._ocultar_sugerencias()
        self.input_box.focus()

    def _saludo_inicial(self):
        if self.estado == "editando":
            nombre = os.path.splitext(os.path.basename(self.ultimo_path))[0]
            self._agregar_burbuja(f"Editando '{nombre}'. Que queres modificar?")
        elif self.prompt_inicial:
            tipo_str = "Excel" if self.tipo == "excel" else "Word"
            self._agregar_burbuja(f"Generando {tipo_str} con la plantilla seleccionada...")
            self._agregar_burbuja(self.prompt_inicial, es_ia=False)
            self._set_input(False)
            threading.Thread(target=self._generar, args=(self.prompt_inicial,), daemon=True).start()
        else:
            tipo_str = "Excel" if self.tipo == "excel" else "Word"
            self._agregar_burbuja(f"Hola! Describe el {tipo_str} que queres generar.")

    def _enviar(self):
        texto = self.input_box.get().strip()
        if not texto:
            return
        self.input_box.delete(0, "end")
        self._agregar_burbuja(texto, es_ia=False)
        self._set_input(False)

        self.historial_mensajes.append(texto)
        if len(self.historial_mensajes) > 3:
            self.historial_mensajes.pop(0)

        if self.estado == "esperando_prompt":
            self._agregar_burbuja("Procesando tu pedido...")
            threading.Thread(target=self._generar, args=(texto,), daemon=True).start()
        elif self.estado == "esperando_nombre":
            self._guardar_archivo(texto)
        elif self.estado == "editando":
            self._agregar_burbuja("Aplicando cambios...")
            threading.Thread(target=self._editar, args=(texto,), daemon=True).start()

    def _generar(self, prompt):
        try:
            from back.ai import mejorar_prompt, buscar_datos_web, generacion_json, parsear_json, analizar_datos_web
            from back.config import instrucciones_excel, instrucciones_word

            imagenes, prompt_limpio = self._extraer_imagenes(prompt)
            contexto_hist = self._construir_contexto_historial()

            contexto_empresa = get_contexto_empresa(self.master.sesion)

            self._agregar_burbuja("Mejorando tu prompt...")
            prompt_mejorado = mejorar_prompt(prompt_limpio, self.seleccion, contexto_empresa)

            if contexto_hist:
                prompt_mejorado = contexto_hist + prompt_mejorado

            if self.seleccion == "1":
                self._agregar_burbuja("Buscando datos en la web...")
                datos_web = buscar_datos_web(prompt_mejorado)
                self._agregar_burbuja("Analizando datos encontrados...")
                datos_analizados = analizar_datos_web(prompt_mejorado, datos_web, "Excel")
                prompt_final = f"{prompt_mejorado}\n\nVerified real data:\n{datos_analizados}"
            else:
                prompt_final = prompt_mejorado

            instrucciones = instrucciones_excel if self.seleccion == "1" else instrucciones_word
            contenido = generacion_json(prompt_final, instrucciones)

            data = parsear_json(contenido)
            if data is None:
                self._agregar_burbuja("No pude generar un resultado valido. Intenta de nuevo.")
                self._set_input(True)
                return

            if imagenes and self.seleccion == "2":
                data["imagenes"] = imagenes
                detalles = ", ".join([f"{img['nombre']} ({img['escala']}cm)" for img in imagenes])
                self._agregar_burbuja(f"Imagenes anadidas: {detalles}")

            if self.indice_pendiente:
                data["indice"] = self.indice_pendiente

            self.ultimo_json = data
            self.estado = "esperando_nombre"
            self._agregar_burbuja("Listo! Que nombre le pones al archivo? (sin extension)")
            self._set_input(True)

        except Exception as e:
            self._agregar_burbuja(f"Error: {str(e)}")
            self._set_input(True)

    def _guardar_archivo(self, nombre):
        try:
            import pandas as pd
            from back.excel import formatear_excel, aplicar_formulas
            from back.word import generar_word
            from openpyxl import load_workbook

            os.makedirs(EXCELS_DIR if self.seleccion == "1" else WORDS_DIR, exist_ok=True)

            empresa = (self.master.sesion or {}).get("empresa", {})

            if self.seleccion == "1":
                path = os.path.join(EXCELS_DIR, nombre + ".xlsx")
                df = pd.DataFrame(self.ultimo_json["datos"], columns=self.ultimo_json["columnas"])
                df.to_excel(path, index=False)
                formatear_excel(path, self.ultimo_json.get("estilo", {}))
                wb = load_workbook(path)
                aplicar_formulas(wb.active)
                wb.save(path)
                wb.close()
            else:
                path = os.path.join(WORDS_DIR, nombre + ".docx")
                generar_word(self.ultimo_json, path, empresa=empresa if empresa_completa(empresa) else None)

            guardar_json_documento(path, self.ultimo_json)
            self.ultimo_path = path

            self._agregar_burbuja("Archivo guardado. Revisa el resultado antes de usarlo.")
            self._agregar_boton_abrir(path)
            self._agregar_burbuja("Podes recibirlo por mail desde la Biblioteca.")
            self._agregar_burbuja("Queres modificar algo? Describe los cambios o escribe 'no' para terminar.")
            self.estado = "editando"
            self._set_input(True)

            sesion = self.master.sesion or {}
            if sesion.get("codigo"):
                threading.Thread(
                    target=self._sync_firebase, args=(path, nombre), daemon=True
                ).start()

        except Exception as e:
            self._agregar_burbuja(f"Error al guardar: {str(e)}")
            self._set_input(True)

    def _sync_firebase(self, path, nombre):
        try:
            from back.firebase import guardar_documento
            sesion = self.master.sesion or {}
            codigo = sesion.get("codigo", "")
            if not codigo:
                return
            with open(path, "rb") as fh:
                archivo_bytes = fh.read()
            tipo = "excel" if self.seleccion == "1" else "word"

            json_path = os.path.splitext(path)[0] + ".json"
            doc_id_existente = None
            if os.path.exists(json_path):
                try:
                    with open(json_path, "r", encoding="utf-8") as fh:
                        data_local = json.load(fh)
                    doc_id_existente = data_local.get("_firebase_doc_id")
                except Exception:
                    pass

            doc_id = guardar_documento(codigo, nombre, tipo, self.ultimo_json, archivo_bytes, doc_id=doc_id_existente)

            if doc_id and os.path.exists(json_path):
                try:
                    with open(json_path, "r", encoding="utf-8") as fh:
                        data_local = json.load(fh)
                    data_local["_firebase_doc_id"] = doc_id
                    with open(json_path, "w", encoding="utf-8") as fh:
                        json.dump(data_local, fh, ensure_ascii=False, indent=2)
                except Exception:
                    pass
        except Exception as e:
            print("Firebase sync error:", e)

    def _editar(self, pedido):
        from back.ai import interpretar_fin

        intencion = interpretar_fin(pedido)
        if intencion == "no":
            self._agregar_burbuja("Perfecto. Podes volver al menu cuando quieras.")
            self._set_input(True)
            return

        imagenes, pedido_limpio = self._extraer_imagenes(pedido)
        contexto = self._construir_contexto_historial()

        try:
            from back.ai import editar_json, parsear_json
            from back.config import instrucciones_excel, instrucciones_word
            from back.excel import formatear_excel, aplicar_formulas
            from openpyxl import load_workbook
            from back.word import generar_word
            import pandas as pd

            instrucciones = instrucciones_excel if self.seleccion == "1" else instrucciones_word

            pedido_con_contexto = pedido_limpio
            if contexto:
                pedido_con_contexto = contexto + pedido_limpio

            contenido = editar_json(
                json.dumps(self.ultimo_json, ensure_ascii=False),
                pedido_con_contexto, instrucciones
            )

            data = parsear_json(contenido)
            if data is None:
                self._agregar_burbuja("No pude aplicar los cambios. Intenta describirlos de otra forma.")
                self._set_input(True)
                return

            if imagenes and self.seleccion == "2":
                data["imagenes"] = imagenes
                detalles = ", ".join([f"{img['nombre']} ({img['escala']}cm)" for img in imagenes])
                self._agregar_burbuja(f"Imagenes anadidas: {detalles}")

            if self.indice_pendiente:
                data["indice"] = self.indice_pendiente

            self.ultimo_json = data
            path = self.ultimo_path

            empresa = (self.master.sesion or {}).get("empresa", {})

            if path:
                if self.seleccion == "1":
                    df = pd.DataFrame(data["datos"], columns=data["columnas"])
                    df.to_excel(path, index=False)
                    formatear_excel(path, data.get("estilo", {}))
                    wb = load_workbook(path)
                    aplicar_formulas(wb.active)
                    wb.save(path)
                    wb.close()
                else:
                    generar_word(data, path, empresa=empresa if empresa_completa(empresa) else None)

                guardar_json_documento(path, data)

            self._agregar_burbuja("Cambios aplicados. Algo mas?")
            self._agregar_boton_abrir(path)
            self._set_input(True)

            sesion = self.master.sesion or {}
            if sesion.get("codigo") and path:
                nombre = os.path.splitext(os.path.basename(path))[0]
                threading.Thread(
                    target=self._sync_firebase, args=(path, nombre), daemon=True
                ).start()

        except Exception as e:
            self._agregar_burbuja(f"Error: {str(e)}")
            self._set_input(True)


class BibliotecaFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color=COLORS["bg"])
        self.master = master
        self.docs_firebase = []
        self._build_ui()

    def _build_ui(self):
        header = ctk.CTkFrame(self, fg_color=COLORS["surface"], height=72, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkButton(
            header, text="←", width=40, height=36,
            fg_color="transparent", hover_color=COLORS["surface2"],
            text_color=COLORS["text_dim"],
            command=self.master.mostrar_menu, font=f(18, True)
        ).pack(side="left", padx=8, pady=8)

        ctk.CTkLabel(header, text="Biblioteca",
                     font=f(15, True), text_color=COLORS["text"]).pack(side="left", padx=4)

        ctk.CTkButton(
            header, text="Sincronizar",
            width=100, height=34,
            fg_color=COLORS["surface2"], hover_color=COLORS["accent"],
            text_color=COLORS["text"], font=f(11, True), corner_radius=10,
            command=self._sincronizar
        ).pack(side="right", padx=12, pady=8)

        self.scroll = ctk.CTkScrollableFrame(self, fg_color=COLORS["bg"])
        self.scroll.pack(fill="both", expand=True, padx=16, pady=16)

        self._cargar_entradas()

        sesion = self.master.sesion or {}
        if sesion.get("codigo"):
            threading.Thread(target=self._cargar_firebase, daemon=True).start()

    def _cargar_entradas(self, firebase_extra=None):
        for widget in self.scroll.winfo_children():
            widget.destroy()

        entries = escanear_biblioteca()
        nombres_locales = {e["nombre"] for e in entries}

        if firebase_extra:
            for doc in firebase_extra:
                if doc["nombre"] not in nombres_locales:
                    entries.append({
                        "nombre":    doc["nombre"],
                        "tipo":      doc["tipo"],
                        "fecha":     doc["fecha"],
                        "path":      None,
                        "json_path": None,
                        "doc_id":    doc["doc_id"],
                    })

        if not entries:
            ctk.CTkLabel(
                self.scroll,
                text="No hay archivos generados todavia.",
                text_color=COLORS["text_dim"], font=f(14)
            ).pack(pady=40)
            return

        for entry in entries:
            self._agregar_entrada(entry)

    def _cargar_firebase(self):
        from back.firebase import cargar_documentos
        sesion = self.master.sesion or {}
        codigo = sesion.get("codigo", "")
        if not codigo:
            return
        docs = cargar_documentos(codigo)
        self.docs_firebase = docs
        self.after(0, lambda: self._cargar_entradas(firebase_extra=docs))

    def _sincronizar(self):
        sesion = self.master.sesion or {}
        if sesion.get("codigo"):
            threading.Thread(target=self._cargar_firebase, daemon=True).start()

    def _agregar_entrada(self, entry):
        row = ctk.CTkFrame(self.scroll, fg_color=COLORS["surface"], corner_radius=14)
        row.pack(fill="x", pady=5)

        icono = "📊" if entry["tipo"] == "excel" else "📄"
        es_local = entry.get("path") is not None

        ctk.CTkLabel(
            row, text=f"{icono}  {entry['nombre']}",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text"], anchor="w"
        ).pack(side="left", padx=14, pady=12)

        ctk.CTkLabel(
            row, text=entry["fecha"],
            font=f(12), text_color=COLORS["text_dim"]
        ).pack(side="left", padx=8)

        if not es_local:
            ctk.CTkLabel(
                row, text="nube",
                font=f(10), text_color=COLORS["accent"]
            ).pack(side="left", padx=4)

        ctk.CTkButton(
            row, text="Borrar",
            width=54, height=30,
            fg_color=COLORS["surface2"], hover_color=COLORS["error"],
            text_color=COLORS["text_dim"], font=f(12), corner_radius=12,
            command=lambda e=entry, r=row: self._borrar(e, r)
        ).pack(side="right", padx=(0, 8), pady=12)

        if es_local:
            ctk.CTkButton(
                row, text="Editar",
                width=60, height=30,
                fg_color=COLORS["surface2"], hover_color=COLORS["accent"],
                text_color=COLORS["text"], font=f(12, True), corner_radius=12,
                command=lambda e=entry: self._editar(e)
            ).pack(side="right", padx=(0, 6), pady=12)

            ctk.CTkButton(
                row, text="Abrir",
                width=56, height=30,
                fg_color=COLORS["accent"], hover_color=COLORS["accent_dim"],
                text_color="#000", font=f(12, True), corner_radius=12,
                command=lambda p=entry["path"]: os.startfile(p) if os.path.exists(p) else None
            ).pack(side="right", padx=(0, 6), pady=12)

            ctk.CTkButton(
                row, text="Mail",
                width=50, height=30,
                fg_color=COLORS["surface2"], hover_color=COLORS["accent"],
                text_color=COLORS["text"], font=f(12, True), corner_radius=12,
                command=lambda e=entry: self._abrir_mail_popup(e)
            ).pack(side="right", padx=(0, 6), pady=12)
        else:
            ctk.CTkButton(
                row, text="Descargar",
                width=86, height=30,
                fg_color=COLORS["accent"], hover_color=COLORS["accent_dim"],
                text_color="#000", font=f(12, True), corner_radius=12,
                command=lambda e=entry: self._descargar(e)
            ).pack(side="right", padx=(0, 6), pady=12)

    def _borrar(self, entry, row):
        if entry.get("path") and os.path.exists(entry["path"]):
            os.remove(entry["path"])
        if entry.get("json_path") and os.path.exists(entry["json_path"]):
            os.remove(entry["json_path"])

        sesion = self.master.sesion or {}
        if entry.get("doc_id") and sesion.get("codigo"):
            def _del():
                from back.firebase import eliminar_documento
                eliminar_documento(sesion["codigo"], entry["doc_id"])
            threading.Thread(target=_del, daemon=True).start()

        row.destroy()

    def _editar(self, entry):
        data = cargar_json_documento(entry["path"])
        if data is None:
            return
        self.master.mostrar_chat(
            tipo=entry["tipo"],
            json_inicial=data,
            path_inicial=entry["path"]
        )

    def _descargar(self, entry):
        sesion = self.master.sesion or {}
        codigo = sesion.get("codigo", "")
        if not codigo or not entry.get("doc_id"):
            return

        def _worker():
            from back.firebase import descargar_documento
            archivo_bytes = descargar_documento(codigo, entry["doc_id"])
            if not archivo_bytes:
                return

            ext = ".xlsx" if entry["tipo"] == "excel" else ".docx"
            carpeta = EXCELS_DIR if entry["tipo"] == "excel" else WORDS_DIR
            os.makedirs(carpeta, exist_ok=True)
            path = os.path.join(carpeta, entry["nombre"] + ext)

            with open(path, "wb") as fh:
                fh.write(archivo_bytes)

            if entry.get("json_data"):
                guardar_json_documento(path, entry["json_data"])

            self.after(0, lambda: os.startfile(path))
            self.after(100, self._cargar_entradas)

        threading.Thread(target=_worker, daemon=True).start()

    def _abrir_mail_popup(self, entry):
        top = ctk.CTkToplevel(self)
        top.title("Enviar archivo")
        top.geometry("350x160")
        top.configure(fg_color=COLORS["bg"])

        ctk.CTkLabel(
            top, text="Ingresa tu mail para recibir el archivo",
            text_color=COLORS["text"]
        ).pack(pady=(15, 5))

        entry_mail = ctk.CTkEntry(
            top, width=250,
            fg_color=COLORS["surface2"], text_color=COLORS["text"]
        )
        entry_mail.pack(pady=5)

        error_label = ctk.CTkLabel(top, text="", text_color=COLORS["error"])
        error_label.pack()

        def enviar():
            mail = entry_mail.get().strip()
            if not re.match(r"^[^@]+@[^@]+\.[^@]+$", mail):
                error_label.configure(text="Mail invalido")
                return
            try:
                enviar_mail(mail, entry["path"])
                top.destroy()
            except Exception:
                error_label.configure(text="Error al enviar")

        ctk.CTkButton(
            top, text="Enviar",
            fg_color=COLORS["accent"], text_color="#000",
            command=enviar
        ).pack(pady=10)


class GaleriaFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color=COLORS["bg"])
        self.master = master
        self.galeria = cargar_galeria()
        self._build_ui()

    def _build_ui(self):
        header = ctk.CTkFrame(self, fg_color=COLORS["surface"], height=72, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkButton(
            header, text="←", width=40, height=36,
            fg_color="transparent", hover_color=COLORS["surface2"],
            text_color=COLORS["text_dim"],
            command=self.master.mostrar_menu, font=f(18, True)
        ).pack(side="left", padx=8, pady=8)

        ctk.CTkLabel(header, text="Galeria de Imagenes",
                     font=f(15, True), text_color=COLORS["text"]).pack(side="left", padx=4)

        ctk.CTkButton(
            header, text="+ Subir",
            width=80, height=36,
            fg_color=COLORS["accent"], hover_color=COLORS["accent_dim"],
            text_color="#000", font=ctk.CTkFont(size=12, weight="bold"),
            command=self._cargar_imagen
        ).pack(side="right", padx=(0, 8), pady=8)

        self.scroll = ctk.CTkScrollableFrame(self, fg_color=COLORS["bg"])
        self.scroll.pack(fill="both", expand=True, padx=16, pady=16)

        self._cargar_imagenes()

    def _cargar_imagenes(self):
        for widget in self.scroll.winfo_children():
            widget.destroy()

        if not self.galeria:
            ctk.CTkLabel(
                self.scroll, text="No hay imagenes cargadas. Subi una!",
                text_color=COLORS["text_dim"], font=f(14)
            ).pack(pady=40)
            return

        for nombre, datos in self.galeria.items():
            self._agregar_imagen(nombre, datos)

    def _agregar_imagen(self, nombre, datos):
        row = ctk.CTkFrame(
            self.scroll, fg_color=COLORS["surface"],
            corner_radius=16, border_width=1, border_color=COLORS["border"]
        )
        row.pack(fill="x", pady=8, padx=4)

        try:
            img = PIL.Image.open(datos["path"])
            img.thumbnail((90, 90))
            preview = ctk.CTkImage(light_image=img, dark_image=img, size=(90, 90))
            img_label = ctk.CTkLabel(row, text="", image=preview)
            img_label.image = preview
            img_label.pack(side="left", padx=12, pady=12)
        except Exception:
            ctk.CTkLabel(row, text="", font=f(35)).pack(side="left", padx=20)

        info = ctk.CTkFrame(row, fg_color="transparent")
        info.pack(side="left", fill="both", expand=True, padx=10)

        ctk.CTkLabel(info, text=nombre, font=f(15, True),
                     text_color=COLORS["text"], anchor="w").pack(anchor="w", pady=(8, 0))

        ctk.CTkLabel(info, text=f"/{nombre}", font=f(12),
                     text_color=COLORS["accent"]).pack(anchor="w")

        botones = ctk.CTkFrame(row, fg_color="transparent")
        botones.pack(side="right", padx=10)

        ctk.CTkButton(
            botones, text="Copiar",
            width=60, height=30, font=f(11),
            fg_color=COLORS["surface2"], hover_color=COLORS["accent"],
            text_color=COLORS["text"], corner_radius=8,
            command=lambda: [self.clipboard_clear(), self.clipboard_append(f"/{nombre}"), self.update()]
        ).pack(pady=4)

        ctk.CTkButton(
            botones, text="Borrar",
            width=60, height=30, font=f(11),
            fg_color=COLORS["error"], hover_color="#CC3344",
            text_color="#fff", corner_radius=8,
            command=lambda n=nombre: self._borrar_imagen(n, row)
        ).pack(pady=4)

    def _cargar_imagen(self):
        from tkinter import filedialog
        import shutil

        ruta = filedialog.askopenfilename(
            title="Selecciona una imagen",
            filetypes=[("Imagenes", "*.jpg *.jpeg *.png *.bmp"), ("Todos", "*.*")]
        )
        if not ruta:
            return

        try:
            try:
                from PIL import Image
                img = Image.open(ruta)
                img.verify()
            except ImportError:
                pass
            except Exception:
                return

            nombre = os.path.splitext(os.path.basename(ruta))[0]
            nombre_original = nombre
            contador = 1
            while nombre in self.galeria:
                nombre = f"{nombre_original}_{contador}"
                contador += 1

            os.makedirs(GALLERY_DIR, exist_ok=True)
            ext = os.path.splitext(ruta)[1]
            path_dest = os.path.join(GALLERY_DIR, nombre + ext)
            shutil.copy2(ruta, path_dest)

            self.galeria[nombre] = {
                "path": path_dest,
                "fecha": datetime.now().strftime("%Y-%m-%d %H:%M")
            }
            guardar_galeria(self.galeria)
            self._cargar_imagenes()

        except Exception as e:
            print(f"Error al cargar imagen: {e}")

    def _borrar_imagen(self, nombre, row):
        if nombre in self.galeria:
            path = self.galeria[nombre]["path"]
            if os.path.exists(path):
                os.remove(path)
            del self.galeria[nombre]
            guardar_galeria(self.galeria)
            row.destroy()


if __name__ == "__main__":
    app = App()
    app.mainloop()
