from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

_KEYWORDS_PRECIO    = ["precio", "price", "unitario", "costo", "valor", "subtotal", "iva", "total", "importe"]
_KEYWORDS_FECHA     = ["fecha", "date", "alta", "emision", "vencimiento", "ingreso"]
_KEYWORDS_TOTAL     = ["total", "subtotal"]
_KEYWORDS_PRECIO_U  = ["precio", "price", "unitario", "costo", "valor"]
_KEYWORDS_CANTIDAD  = ["cantidad", "quantity", "vendida", "stock", "unidades"]
_KEYWORDS_IVA       = ["iva", "impuesto", "tax"]
_KEYWORDS_PROMEDIO  = ["promedio", "average"]


def _buscar_col(headers, keywords):
    for i, h in enumerate(headers):
        if h and any(k.lower() in h.lower() for k in keywords):
            return i + 1
    return None


def formatear_excel(path, estilo):
    wb = load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    header_fill = PatternFill("solid", fgColor=estilo.get("header_color", "2F4F7F"))
    header_font = Font(
        bold=True,
        color=estilo.get("font_color_header", "FFFFFF"),
        size=estilo.get("font_size", 11)
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    cols_precio = [
        i + 1 for i, h in enumerate(headers)
        if h and any(k.lower() in h.lower() for k in _KEYWORDS_PRECIO)
    ]
    cols_fecha = [
        i + 1 for i, h in enumerate(headers)
        if h and any(k.lower() in h.lower() for k in _KEYWORDS_FECHA)
    ]

    fill_alt = PatternFill("solid", fgColor=estilo.get("row_alt_color", "F2F2F2"))
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            if i % 2 == 0:
                cell.fill = fill_alt
            cell.font = Font(size=estilo.get("font_size", 11))
            cell.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")

            if cell.column in cols_precio:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="top")

            if cell.column in cols_fecha:
                cell.number_format = 'DD/MM/YYYY'
                cell.alignment = Alignment(horizontal="center", vertical="top")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    wb.save(path)
    wb.close()


def aplicar_formulas(ws):
    headers = [cell.value for cell in ws[1]]

    col_total    = _buscar_col(headers, _KEYWORDS_TOTAL)
    col_subtotal = _buscar_col(headers, ["subtotal"])
    col_precio   = _buscar_col(headers, _KEYWORDS_PRECIO_U)
    col_cantidad = _buscar_col(headers, _KEYWORDS_CANTIDAD)
    col_iva      = _buscar_col(headers, _KEYWORDS_IVA)
    col_promedio = _buscar_col(headers, _KEYWORDS_PROMEDIO)

    ultima_fila = ws.max_row

    for row in range(2, ultima_fila + 1):
        if col_subtotal and col_precio and col_cantidad:
            p = ws.cell(row=row, column=col_precio).coordinate
            q = ws.cell(row=row, column=col_cantidad).coordinate
            ws.cell(row=row, column=col_subtotal).value = f"={p}*{q}"

        if col_iva and col_subtotal:
            s = ws.cell(row=row, column=col_subtotal).coordinate
            ws.cell(row=row, column=col_iva).value = f"={s}*0.21"
        elif col_iva and col_precio and col_cantidad:
            p = ws.cell(row=row, column=col_precio).coordinate
            q = ws.cell(row=row, column=col_cantidad).coordinate
            ws.cell(row=row, column=col_iva).value = f"={p}*{q}*0.21"

        if col_total:
            if col_subtotal and col_iva:
                s = ws.cell(row=row, column=col_subtotal).coordinate
                v = ws.cell(row=row, column=col_iva).coordinate
                ws.cell(row=row, column=col_total).value = f"={s}+{v}"
            elif col_precio and col_cantidad:
                p = ws.cell(row=row, column=col_precio).coordinate
                q = ws.cell(row=row, column=col_cantidad).coordinate
                ws.cell(row=row, column=col_total).value = f"={p}*{q}"

    fila_suma = ultima_fila + 2

    bold_font = Font(bold=True, size=11)

    for col_idx, label in [
        (col_subtotal, "SUBTOTAL"),
        (col_iva,      "IVA TOTAL"),
        (col_total,    "TOTAL"),
        (col_cantidad, None),
        (col_promedio, "PROMEDIO"),
    ]:
        if not col_idx:
            continue
        letra = get_column_letter(col_idx)

        if col_idx == col_promedio:
            ws.cell(row=fila_suma, column=col_idx).value = f"=AVERAGE({letra}2:{letra}{ultima_fila})"
        else:
            ws.cell(row=fila_suma, column=col_idx).value = f"=SUM({letra}2:{letra}{ultima_fila})"

        ws.cell(row=fila_suma, column=col_idx).font = bold_font
        ws.cell(row=fila_suma, column=col_idx).number_format = '$#,##0.00'

        if label and col_idx > 1:
            label_cell = ws.cell(row=fila_suma, column=col_idx - 1)
            label_cell.value = label
            label_cell.font = bold_font
